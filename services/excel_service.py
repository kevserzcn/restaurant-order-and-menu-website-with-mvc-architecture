import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils.datetime_utils import format_local
import os

def generate_report_excel(payments):
    
    timestamp = format_local(datetime.utcnow(), '%Y%m%d_%H%M%S')
    filename = f"rapor_{timestamp}.xlsx"
    filepath = os.path.join('static', 'reports', filename)
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gelir Raporu"
    
    ws['A1'] = "ABLALARIN YERİ - GELİR RAPORU"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Rapor Tarihi: {format_local(datetime.utcnow())}"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:H2')
    
    total_amount = sum(p.amount for p in payments)
    ws['A3'] = f"Toplam Gelir: {total_amount:.2f} ₺"
    ws['A3'].font = Font(size=12, bold=True)
    ws.merge_cells('A3:H3')
    
    ws['A4'] = ""
    
    headers = ['Ödeme ID', 'Sipariş ID', 'Masa', 'Müşteri', 'Tutar', 'Yöntem', 'Tarih']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    for row, payment in enumerate(payments, 6):
        order = payment.order
        table_number = 'N/A'
        user_name = 'N/A'
        
        if order:
            if hasattr(order, 'table') and order.table:
                table_number = order.table.name
            if hasattr(order, 'user') and order.user:
                user_name = order.user.name
        
        data = [
            payment.id,
            payment.order_id,
            table_number,
            user_name,
            f"{payment.amount:.2f} ₺",
            payment.payment_method.title(),
            format_local(payment.created_at)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    column_widths = [12, 12, 10, 20, 15, 15, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    wb.save(filepath)
    
    return filepath
